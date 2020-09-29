from geopy.geocoders import Nominatim
geolocator = Nominatim(user_agent="my-application", timeout = 20)

zipcode1 = "Quality Inn Palm Springs Downtown, Palm Springs"
print("\nZipcode:",zipcode1)
location = geolocator.geocode(zipcode1, country_codes= "US")
print("Details of the said pincode:")
print(location.latitude, location.longitude)


postcode_col_num = 1
country_code_col_num = 2
lat_col_num = 3
long_col_num = 4



count = 0
n = 0

import openpyxl
my_folder_path = "/Users/rossmcintyre/Documents/Datasets/"
file_name = "Active_Ambassador_Locations.xlsx"
my_path = my_folder_path + file_name
my_wb_obj = openpyxl.load_workbook(my_path)

my_sheet_obj = my_wb_obj['Sheet6']


for i in range(192, my_sheet_obj.max_row + 1):
    my_cell_obj = my_sheet_obj.cell(row = i, column = postcode_col_num)
    
    

    if my_cell_obj.value:

        postcode = my_cell_obj.value
        print(postcode)
        location = geolocator.geocode(postcode, country_codes = str(my_sheet_obj.cell(row = i, column = country_code_col_num)))

        try:     
            
            latitude = my_sheet_obj.cell(row = i, column = lat_col_num)
            latitude.value = str(location.latitude)
            
            
            longitude = my_sheet_obj.cell(row = i, column = long_col_num)
            longitude.value = str(location.longitude)

            count += 1
            if count % 100 == 0:
                my_wb_obj.save(my_path)
                print("Saved, %d done, only %d to go" % (count, my_sheet_obj.max_row - i))

            
        except: 
            latitude = my_sheet_obj.cell(row = i, column = lat_col_num)
            latitude.value = "Error"

            longitude = my_sheet_obj.cell(row = i, column = long_col_num)
            longitude.value = "Error"

                
            count += 1
            continue
    
    else:
        continue

my_wb_obj.save(my_path)
